from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError

from app.suppliers.schemas import (
    AliasInput,
    ImportIssue,
    ProductInput,
    SiteInput,
    SupplierCreate,
    clean_optional_text,
    normalize_alias,
)

MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ROWS_PER_SHEET = 5_000
MAX_ISSUES = 200

SHEET_SUPPLIERS = "供应商"
SHEET_SITES = "生产地点"
SHEET_PRODUCTS = "供应产品"

SUPPLIER_COLUMNS = [
    ("供应商编码", "supplier_code", True, "我司内部唯一编码，例如 SUP-0001"),
    ("法人主体", "legal_name", True, "供应商法人主体全称"),
    ("国家代码", "country_code", True, "ISO 两位国家代码，例如 CN、US"),
    ("注册编号", "registry_no", False, "统一社会信用代码或境外注册编号"),
    ("注册地址", "registration_address", False, "供应商工商登记的详细地址"),
    ("行业", "industry", False, "行业标签，例如 电梯配件、钣金加工；用于宏观行业匹配"),
    (
        "关键原材料",
        "raw_materials",
        False,
        "多个关键原材料或上游依赖使用英文分号分隔；用于行业/原材料匹配",
    ),
    ("别名", "aliases", False, "多个中文名、英文名或简称使用英文分号分隔"),
    ("启用监控", "enabled", False, "TRUE 或 FALSE，留空时默认为 TRUE"),
]
SITE_COLUMNS = [
    ("供应商编码", "supplier_code", True, "必须存在于“供应商”工作表"),
    ("地点名称", "site_name", True, "工厂、仓库或其他实际履约地点名称"),
    ("国家代码", "country_code", True, "ISO 两位国家代码，例如 CN、US"),
    ("省州地区", "region", False, "省、州或一级行政区"),
    ("城市", "city", False, "城市"),
    ("区县", "district", False, "区、县、旗或同级行政区"),
    ("详细地址", "address", True, "实际生产或履约地址"),
    ("纬度", "latitude", False, "-90 到 90；必须与经度同时填写"),
    ("经度", "longitude", False, "-180 到 180；必须与纬度同时填写"),
]
PRODUCT_COLUMNS = [
    ("供应商编码", "supplier_code", True, "必须存在于“供应商”工作表"),
    ("供应产品", "name", True, "供应产品、物料类别或服务名称"),
    ("关键词", "keywords", False, "多个关键词使用英文分号分隔"),
]


@dataclass
class ImportPlan:
    suppliers: list[SupplierCreate]


class WorkbookValidationError(ValueError):
    def __init__(self, issues: list[ImportIssue]) -> None:
        super().__init__("供应商工作簿校验失败")
        self.issues = issues


def create_template() -> bytes:
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)
    _add_template_sheet(
        workbook,
        SHEET_SUPPLIERS,
        SUPPLIER_COLUMNS,
        [20, 34, 14, 24, 34, 20, 26, 36, 14],
    )
    _add_template_sheet(
        workbook, SHEET_SITES, SITE_COLUMNS, [20, 24, 14, 20, 18, 18, 42, 14, 14]
    )
    _add_template_sheet(workbook, SHEET_PRODUCTS, PRODUCT_COLUMNS, [20, 28, 42])

    enabled_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    workbook[SHEET_SUPPLIERS].add_data_validation(enabled_validation)
    enabled_validation.add("I2:I5000")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _add_template_sheet(
    workbook: Workbook,
    title: str,
    columns: list[tuple[str, str, bool, str]],
    widths: list[int],
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{chr(64 + len(columns))}1"
    sheet.row_dimensions[1].height = 28

    for index, ((label, field, required, description), width) in enumerate(
        zip(columns, widths, strict=True), start=1
    ):
        cell = sheet.cell(row=1, column=index, value=f"{label}{' *' if required else ''}")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(description, "JoeMu")
        sheet.column_dimensions[cell.column_letter].width = width
        if field in {"supplier_code", "registry_no"}:
            sheet.column_dimensions[cell.column_letter].number_format = "@"


def parse_workbook(data: bytes) -> ImportPlan:
    issues: list[ImportIssue] = []
    _validate_archive(data, issues)
    if issues:
        raise WorkbookValidationError(issues)

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    except (BadZipFile, OSError, ValueError) as exc:
        raise WorkbookValidationError(
            [ImportIssue(sheet="工作簿", message=f"无法读取 .xlsx 文件：{exc}")]
        ) from exc

    expected_sheets = {SHEET_SUPPLIERS, SHEET_SITES, SHEET_PRODUCTS}
    missing = expected_sheets.difference(workbook.sheetnames)
    for sheet_name in sorted(missing):
        _add_issue(issues, ImportIssue(sheet=sheet_name, message="缺少必需工作表"))
    if issues:
        raise WorkbookValidationError(issues)

    supplier_rows = _read_rows(workbook[SHEET_SUPPLIERS], SUPPLIER_COLUMNS, issues)
    site_rows = _read_rows(workbook[SHEET_SITES], SITE_COLUMNS, issues)
    product_rows = _read_rows(workbook[SHEET_PRODUCTS], PRODUCT_COLUMNS, issues)

    builders: dict[str, dict[str, object]] = {}
    supplier_source_rows: dict[str, int] = {}
    registry_owners: dict[tuple[str, str], str] = {}

    for row_number, row in supplier_rows:
        code = _required_text(row.get("supplier_code"))
        if not code:
            _add_issue(
                issues,
                ImportIssue(
                    sheet=SHEET_SUPPLIERS,
                    row=row_number,
                    field="供应商编码",
                    message="不能为空",
                ),
            )
            continue
        if code in builders:
            _add_issue(
                issues,
                ImportIssue(
                    sheet=SHEET_SUPPLIERS,
                    row=row_number,
                    field="供应商编码",
                    message=f"与第 {supplier_source_rows[code]} 行重复",
                ),
            )
            continue

        aliases = [
            AliasInput(alias=value)
            for value in _split_values(clean_optional_text(row.get("aliases")))
        ]
        payload = {
            "supplier_code": code,
            "legal_name": row.get("legal_name"),
            "country_code": row.get("country_code"),
            "registry_no": row.get("registry_no"),
            "registration_address": row.get("registration_address"),
            "industry": clean_optional_text(row.get("industry")),
            "raw_materials": _split_values(clean_optional_text(row.get("raw_materials"))),
            "enabled": _parse_bool(
                row.get("enabled"), SHEET_SUPPLIERS, row_number, "启用监控", issues
            ),
            "aliases": aliases,
            "sites": [],
            "products": [],
        }
        try:
            supplier = SupplierCreate.model_validate(payload)
        except ValidationError as exc:
            _record_validation_errors(exc, SHEET_SUPPLIERS, row_number, issues)
            continue

        if supplier.registry_no:
            registry_key = (supplier.country_code, supplier.registry_no)
            owner = registry_owners.get(registry_key)
            if owner:
                _add_issue(
                    issues,
                    ImportIssue(
                        sheet=SHEET_SUPPLIERS,
                        row=row_number,
                        field="注册编号",
                        message=f"与供应商编码 {owner} 重复",
                    ),
                )
                continue
            registry_owners[registry_key] = code

        builders[code] = supplier.model_dump()
        supplier_source_rows[code] = row_number

    site_keys: set[tuple[str, str]] = set()
    for row_number, row in site_rows:
        code = _required_text(row.get("supplier_code"))
        if code not in builders:
            _add_issue(
                issues,
                ImportIssue(
                    sheet=SHEET_SITES,
                    row=row_number,
                    field="供应商编码",
                    message="未在“供应商”工作表中定义",
                ),
            )
            continue
        site_name = _required_text(row.get("site_name"))
        key = (code, normalize_alias(site_name)) if site_name else None
        if key and key in site_keys:
            _add_issue(
                issues,
                ImportIssue(
                    sheet=SHEET_SITES,
                    row=row_number,
                    field="地点名称",
                    message="同一供应商的地点名称重复",
                ),
            )
            continue
        latitude = _parse_decimal(row.get("latitude"), SHEET_SITES, row_number, "纬度", issues)
        longitude = _parse_decimal(
            row.get("longitude"), SHEET_SITES, row_number, "经度", issues
        )
        try:
            site = SiteInput.model_validate(
                {
                    "site_name": row.get("site_name"),
                    "country_code": row.get("country_code"),
                    "region": row.get("region"),
                    "city": row.get("city"),
                    "district": row.get("district"),
                    "address": row.get("address"),
                    "latitude": latitude,
                    "longitude": longitude,
                }
            )
        except ValidationError as exc:
            _record_validation_errors(exc, SHEET_SITES, row_number, issues)
            continue
        site_keys.add((code, normalize_alias(site.site_name)))
        sites = builders[code]["sites"]
        assert isinstance(sites, list)
        sites.append(site)

    product_keys: set[tuple[str, str]] = set()
    for row_number, row in product_rows:
        code = _required_text(row.get("supplier_code"))
        if code not in builders:
            _add_issue(
                issues,
                ImportIssue(
                    sheet=SHEET_PRODUCTS,
                    row=row_number,
                    field="供应商编码",
                    message="未在“供应商”工作表中定义",
                ),
            )
            continue
        name = _required_text(row.get("name"))
        key = (code, normalize_alias(name)) if name else None
        if key and key in product_keys:
            _add_issue(
                issues,
                ImportIssue(
                    sheet=SHEET_PRODUCTS,
                    row=row_number,
                    field="供应产品",
                    message="同一供应商的产品名称重复",
                ),
            )
            continue
        try:
            product = ProductInput.model_validate(
                {
                    "name": row.get("name"),
                    "keywords": _split_values(clean_optional_text(row.get("keywords"))),
                }
            )
        except ValidationError as exc:
            _record_validation_errors(exc, SHEET_PRODUCTS, row_number, issues)
            continue
        product_keys.add((code, normalize_alias(product.name)))
        products = builders[code]["products"]
        assert isinstance(products, list)
        products.append(product)

    suppliers: list[SupplierCreate] = []
    for code, payload in builders.items():
        sites = payload["sites"]
        products = payload["products"]
        row_number = supplier_source_rows[code]
        if not sites:
            _add_issue(
                issues,
                ImportIssue(
                    sheet=SHEET_SUPPLIERS,
                    row=row_number,
                    field="供应商编码",
                    message="至少需要一个生产地点",
                ),
            )
        if not products:
            _add_issue(
                issues,
                ImportIssue(
                    sheet=SHEET_SUPPLIERS,
                    row=row_number,
                    field="供应商编码",
                    message="至少需要一个供应产品",
                ),
            )
        suppliers.append(SupplierCreate.model_validate(payload))

    if not supplier_rows:
        _add_issue(issues, ImportIssue(sheet=SHEET_SUPPLIERS, message="至少需要一条供应商数据"))
    if issues:
        raise WorkbookValidationError(issues)
    return ImportPlan(suppliers=suppliers)


def _validate_archive(data: bytes, issues: list[ImportIssue]) -> None:
    if not data:
        _add_issue(issues, ImportIssue(sheet="工作簿", message="文件为空"))
        return
    if len(data) > MAX_FILE_BYTES:
        _add_issue(issues, ImportIssue(sheet="工作簿", message="文件不能超过 5MB"))
        return
    if not is_zipfile(BytesIO(data)):
        _add_issue(issues, ImportIssue(sheet="工作簿", message="不是有效的 .xlsx 文件"))
        return
    try:
        with ZipFile(BytesIO(data)) as archive:
            total_size = sum(item.file_size for item in archive.infolist())
    except BadZipFile:
        _add_issue(issues, ImportIssue(sheet="工作簿", message="不是有效的 .xlsx 文件"))
        return
    if total_size > MAX_UNCOMPRESSED_BYTES:
        _add_issue(issues, ImportIssue(sheet="工作簿", message="解压后内容不能超过 50MB"))


def _read_rows(
    sheet: Any,
    columns: list[tuple[str, str, bool, str]],
    issues: list[ImportIssue],
) -> list[tuple[int, dict[str, object]]]:
    max_row = sheet.max_row
    if max_row is not None and max_row > MAX_ROWS_PER_SHEET:
        _add_issue(
            issues,
            ImportIssue(sheet=sheet.title, message="每个工作表最多 5000 行"),
        )
        return []

    expected_labels = [f"{label}{' *' if required else ''}" for label, _, required, _ in columns]
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1), ())
    actual_labels = [clean_optional_text(cell.value) for cell in header_cells[: len(columns)]]
    if actual_labels != expected_labels:
        _add_issue(
            issues,
            ImportIssue(
                sheet=sheet.title,
                row=1,
                message="表头与标准模板不一致，请重新下载模板",
            ),
        )
        return []

    rows: list[tuple[int, dict[str, object]]] = []
    for row_number, cells in enumerate(
        sheet.iter_rows(min_row=2, max_col=len(columns)), start=2
    ):
        if row_number > MAX_ROWS_PER_SHEET:
            _add_issue(
                issues,
                ImportIssue(sheet=sheet.title, message="每个工作表最多 5000 行"),
            )
            break
        if all(cell.value is None or str(cell.value).strip() == "" for cell in cells):
            continue
        if any(cell.data_type == "f" for cell in cells):
            _add_issue(
                issues,
                ImportIssue(
                    sheet=sheet.title,
                    row=row_number,
                    message="不允许使用公式，请粘贴为值",
                ),
            )
            continue
        row = {
            field: cell.value
            for (_label, field, _required, _description), cell in zip(
                columns, cells, strict=True
            )
        }
        rows.append((row_number, row))
    return rows


def _required_text(value: object) -> str:
    return clean_optional_text(value) or ""


def _split_values(value: str | None) -> list[str]:
    if not value:
        return []
    normalized = value.replace("；", ";")
    return list(dict.fromkeys(item.strip() for item in normalized.split(";") if item.strip()))


def _parse_bool(
    value: object,
    sheet: str,
    row: int,
    field: str,
    issues: list[ImportIssue],
) -> bool:
    if value is None or str(value).strip() == "":
        return True
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().casefold()
    if normalized in {"true", "1", "yes", "是"}:
        return True
    if normalized in {"false", "0", "no", "否"}:
        return False
    _add_issue(
        issues,
        ImportIssue(sheet=sheet, row=row, field=field, message="必须是 TRUE 或 FALSE"),
    )
    return True


def _parse_decimal(
    value: object,
    sheet: str,
    row: int,
    field: str,
    issues: list[ImportIssue],
) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        _add_issue(
            issues,
            ImportIssue(sheet=sheet, row=row, field=field, message="必须是有效数字"),
        )
        return None


def _record_validation_errors(
    error: ValidationError,
    sheet: str,
    row: int,
    issues: list[ImportIssue],
) -> None:
    for item in error.errors():
        field = str(item["loc"][-1]) if item["loc"] else None
        _add_issue(
            issues,
            ImportIssue(sheet=sheet, row=row, field=field, message=str(item["msg"])),
        )


def _add_issue(issues: list[ImportIssue], issue: ImportIssue) -> None:
    if len(issues) < MAX_ISSUES:
        issues.append(issue)
