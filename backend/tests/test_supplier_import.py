import re
from collections.abc import Callable
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import load_workbook

from app.suppliers.importer import (
    SHEET_PRODUCTS,
    SHEET_SITES,
    SHEET_SUPPLIERS,
    WorkbookValidationError,
    create_template,
    parse_workbook,
)


def test_template_has_three_structured_sheets() -> None:
    workbook = load_workbook(BytesIO(create_template()))

    assert workbook.sheetnames == [SHEET_SUPPLIERS, SHEET_SITES, SHEET_PRODUCTS]
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref is not None
        assert sheet["A1"].font.bold is True
        assert sheet["A1"].comment is not None
        assert sheet.sheet_view.showGridLines is False
        assert sheet.column_dimensions["A"].number_format == "@"
    assert workbook[SHEET_SUPPLIERS].column_dimensions["D"].number_format == "@"


def test_parse_valid_workbook(workbook_factory: Callable[..., bytes]) -> None:
    plan = parse_workbook(workbook_factory())

    assert len(plan.suppliers) == 1
    supplier = plan.suppliers[0]
    assert supplier.supplier_code == "SUP-0001"
    assert supplier.country_code == "CN"
    assert [item.alias for item in supplier.aliases] == ["测试供应商", "Test Supplier"]
    assert supplier.sites[0].city == "上海市"
    assert supplier.products[0].keywords == ["零部件", "精密加工"]


def test_parse_workbook_without_dimension_metadata(
    workbook_factory: Callable[..., bytes],
) -> None:
    source = BytesIO(workbook_factory())
    output = BytesIO()
    with ZipFile(source) as input_archive, ZipFile(output, "w", ZIP_DEFLATED) as result:
        for item in input_archive.infolist():
            content = input_archive.read(item)
            if item.filename.startswith("xl/worksheets/"):
                content = re.sub(rb"<dimension[^>]*/>", b"", content)
            result.writestr(item, content)

    plan = parse_workbook(output.getvalue())

    assert [supplier.supplier_code for supplier in plan.suppliers] == ["SUP-0001"]


def test_invalid_coordinate_reports_sheet_row_and_field(
    workbook_factory: Callable[..., bytes],
) -> None:
    with pytest.raises(WorkbookValidationError) as captured:
        parse_workbook(workbook_factory(latitude=91))

    issues = captured.value.issues
    assert any(
        issue.sheet == SHEET_SITES
        and issue.row == 2
        and issue.field == "latitude"
        for issue in issues
    )


def test_formula_is_rejected(workbook_factory: Callable[..., bytes]) -> None:
    workbook = load_workbook(BytesIO(workbook_factory()))
    workbook[SHEET_PRODUCTS]["C2"] = "=1+1"
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(WorkbookValidationError) as captured:
        parse_workbook(output.getvalue())

    assert any(issue.message == "不允许使用公式，请粘贴为值" for issue in captured.value.issues)
