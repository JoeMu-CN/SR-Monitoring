from collections.abc import Callable
from io import BytesIO

from fastapi.testclient import TestClient
from httpx import Response
from openpyxl import load_workbook

from app.suppliers.importer import SHEET_SITES


def upload(client: TestClient, content: bytes) -> Response:
    return client.post(
        "/api/v1/suppliers/import",
        files={
            "file": (
                "suppliers.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_template_download(client: TestClient) -> None:
    response = client.get("/api/v1/suppliers/import-template")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert load_workbook(BytesIO(response.content)).sheetnames == [
        "供应商",
        "生产地点",
        "供应产品",
    ]


def test_import_is_idempotent_and_searchable(
    client: TestClient, workbook_factory: Callable[..., bytes]
) -> None:
    first = upload(client, workbook_factory())
    second = upload(client, workbook_factory(legal_name="测试供应商股份有限公司"))

    assert first.status_code == 200
    assert first.json()["created_suppliers"] == 1
    assert second.status_code == 200
    assert second.json()["updated_suppliers"] == 1

    response = client.get("/api/v1/suppliers", params={"q": "Test Supplier", "enabled": True})
    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["legal_name"] == "测试供应商股份有限公司"
    assert len(body["items"][0]["sites"]) == 1
    assert len(body["items"][0]["products"]) == 1


def test_invalid_import_is_atomic(
    client: TestClient, workbook_factory: Callable[..., bytes]
) -> None:
    assert upload(client, workbook_factory()).status_code == 200
    workbook = load_workbook(BytesIO(workbook_factory(supplier_code="SUP-0002")))
    workbook[SHEET_SITES]["G2"] = 100
    output = BytesIO()
    workbook.save(output)

    invalid = upload(client, output.getvalue())

    assert invalid.status_code == 422
    assert invalid.json()["detail"]["errors"][0]["sheet"] == SHEET_SITES
    assert client.get("/api/v1/suppliers").json()["total"] == 1


def test_create_update_and_disable_supplier(client: TestClient) -> None:
    create_payload = {
        "supplier_code": "SUP-API-1",
        "legal_name": "API供应商有限公司",
        "country_code": "cn",
        "registry_no": None,
        "registration_address": "江苏省苏州市姑苏区登记路1号",
        "enabled": True,
        "aliases": [{"alias": "API Supplier", "language": "en"}],
        "sites": [
            {
                "site_name": "苏州工厂",
                "country_code": "CN",
                "region": "江苏省",
                "city": "苏州市",
                "district": "姑苏区",
                "address": "苏州市测试路1号",
                "latitude": 31.2989,
                "longitude": 120.5853,
            }
        ],
        "products": [{"name": "连接器", "keywords": ["电子元件"]}],
    }
    created = client.post("/api/v1/suppliers", json=create_payload)
    assert created.status_code == 201
    supplier_id = created.json()["id"]

    update_payload = {key: value for key, value in create_payload.items() if key != "supplier_code"}
    update_payload["legal_name"] = "API供应商股份有限公司"
    updated = client.put(f"/api/v1/suppliers/{supplier_id}", json=update_payload)
    assert updated.status_code == 200
    assert updated.json()["legal_name"] == "API供应商股份有限公司"
    assert updated.json()["registration_address"] == "江苏省苏州市姑苏区登记路1号"
    assert updated.json()["sites"][0]["district"] == "姑苏区"

    disabled = client.patch(
        f"/api/v1/suppliers/{supplier_id}/enabled",
        json={"enabled": False},
        headers={"X-User-Role": "admin"},
    )
    assert disabled.status_code == 200
    assert disabled.json()["enabled"] is False
    assert client.get("/api/v1/suppliers", params={"enabled": False}).json()["total"] == 1


def test_supplier_keeps_registration_separate_from_multiple_sites(client: TestClient) -> None:
    response = client.post(
        "/api/v1/suppliers",
        json={
            "supplier_code": "SUP-MULTI-SITE",
            "legal_name": "多生产地址供应商",
            "country_code": "CN",
            "registry_no": None,
            "registration_address": "广东省深圳市福田区登记路1号",
            "sites": [
                {
                    "site_name": "深圳工厂",
                    "country_code": "CN",
                    "region": "广东省",
                    "city": "深圳市",
                    "district": "南山区",
                    "address": "广东省深圳市南山区生产路1号",
                },
                {
                    "site_name": "东莞工厂",
                    "country_code": "CN",
                    "region": "广东省",
                    "city": "东莞市",
                    "district": "松山湖",
                    "address": "广东省东莞市松山湖生产路2号",
                },
            ],
            "products": [{"name": "连接器", "keywords": []}],
        },
    )

    assert response.status_code == 201
    body = response.json()
    assert body["registration_address"] == "广东省深圳市福田区登记路1号"
    assert {site["address"] for site in body["sites"]} == {
        "广东省深圳市南山区生产路1号",
        "广东省东莞市松山湖生产路2号",
    }
    assert {site["district"] for site in body["sites"]} == {"南山区", "松山湖"}


def test_supplier_monitoring_toggle_requires_admin(client: TestClient, auth_as) -> None:
    payload = {
        "supplier_code": "SUP-ROLE-1",
        "legal_name": "权限测试供应商",
        "country_code": "CN",
        "registry_no": None,
        "enabled": True,
        "aliases": [],
        "sites": [],
        "products": [],
    }
    created = client.post("/api/v1/suppliers", json=payload)
    supplier_id = created.json()["id"]

    auth_as("viewer", "supplier-viewer")
    viewer = client.patch(
        f"/api/v1/suppliers/{supplier_id}/enabled",
        json={"enabled": False},
        headers={"X-User-Role": "admin"},
    )

    assert viewer.status_code == 403
    assert client.get(f"/api/v1/suppliers/{supplier_id}").json()["enabled"] is True

    auth_as("risk_admin", "supplier-risk-admin")
    admin = client.patch(
        f"/api/v1/suppliers/{supplier_id}/enabled",
        json={"enabled": False},
    )
    assert admin.status_code == 200
    assert admin.json()["enabled"] is False
